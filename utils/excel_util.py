# -- coding: utf-8 --
from openpyxl import load_workbook


class ExcelUtil:
    """
     Excel 文件操作工具类
    """

    _wb = None
    CONFIG = 'Config'

    def __init__(self, filename: str):
        self._wb = load_workbook(filename=filename)

    def getExcelData(self):
        """
        获取 Excel 测试数据
        :return: 用例集合
        """
        case_datas = {}
        # 遍历工作簿
        for sheet in self.sheets():
            title = sheet.title
            if title != self.CONFIG:
                first_row = next(sheet.rows)
                row_index = 2
                max_row = sheet.max_row
                # 取所有 row
                rows = []
                while row_index <= max_row:
                    column_index = 1
                    max_column = sheet.max_column
                    case_data = {}
                    # 取所有 cell
                    while column_index <= max_column:
                        key = first_row[column_index - 1].value
                        value = self.cell(sheet, row_index, column_index)
                        case_data[key] = value
                        column_index += 1
                    rows.append(case_data)
                    row_index += 1
            case_datas[title] = rows
        return case_datas

    # def get_case_number_by_row(self, case_number: str):
    #     case_datas = self.getExcelData()
    #     for


    def getConfig(self):
        """
        获取工作溥配置
        """

        _SHEET_NAME = 'Config'
        config = {}
        for sheet in self.sheets():
            if sheet.title == _SHEET_NAME:
                for row in sheet.rows:
                    key = row[0].value
                    conf_value = row[1].value
                    config[key] = conf_value
                return config
        raise Exception("没有找到工作溥名称为 : %s ,请到文件中添加 " % _SHEET_NAME)

    def sheets(self):
        """
        获取工作溥集合
        """
        return self._wb.worksheets

    def sheet(self, name):
        for sheet in self.sheets():
            if sheet.title == name:
                return sheet
        raise Exception("没有找到工作溥名称为 : %s " % name)

    def rows(self, sheet):
        return sheet.rows

    def cell(self, sheet, r_col, c_col):
        return sheet.cell(row=r_col, column=c_col).value


if __name__ == '__main__':
    excel = ExcelUtil('../test_data/登录接口_test.xlsx')
    datas = excel.getExcelData()
    # datas = excel.get_case_number_by_row('Sheet1/qa_1_value_111')
    print(datas)
